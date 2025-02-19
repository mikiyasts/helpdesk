from email.header import Header
from email.message import EmailMessage
from io import BytesIO
from time import localtime
from django.db.models import Q
from email.mime.text import MIMEText
from django.utils import timezone
from datetime import timedelta
import json
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import timedelta
from django.db.models.functions import Concat
from django.db.models import Count, Avg, F, ExpressionWrapper, fields, Value
from django.utils.dateparse import parse_datetime
from django.contrib.auth.models import Group
from django.utils.http import urlsafe_base64_decode
import uuid
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import render
from django.utils.crypto import get_random_string
import pandas as pd
import requests
from django.conf import settings
from django.contrib.auth.tokens import default_token_generator
from django.contrib.auth import get_user_model
from django.utils.http import urlsafe_base64_encode
from django.core.mail import send_mail
from django.template.loader import render_to_string
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.authentication import BaseAuthentication  
from django.http import Http404
from rest_framework.decorators import api_view
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.exceptions import AuthenticationFailed
from rest_framework.response import Response
import secrets
from .authentication import APIKeyAuthentication
from tickets.models import Attachment, Ticket,TicketCategory, TicketComment, TicketHistory, Acknowledgement
from .serializers import CustomTokenRefreshSerializer, DepartmentSerializer, PasswordResetConfirmSerializer, PasswordResetRequestSerializer, RecentTicketSerializer, ReportTicketSerializer, TicketCategorySerializer, TicketCommentSerializer, TicketHistorySerializer, TicketSerializer, UserCreateSerializer, UserGetSerializer, UserSerializer, AcknowledgementSerializer
from rest_framework.authtoken.models import Token
from rest_framework import status
from users.models import Department, PasswordReset, User,ActivateAccount
from django.shortcuts import get_object_or_404
from rest_framework.decorators import authentication_classes, permission_classes
from rest_framework.authentication import SessionAuthentication, TokenAuthentication
from rest_framework.permissions import IsAuthenticated
import jwt
from rest_framework_simplejwt.views import TokenRefreshView
from rest_framework_simplejwt.serializers import TokenRefreshSerializer
from datetime import datetime
from .models import APIKey
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework_simplejwt.views import TokenObtainPairView
from .serializers import CustomTokenObtainPairSerializer
from django.contrib.auth import get_user_model
from rest_framework import serializers
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from .serializers import JWTUserSerializer
from users.models import User
from .serializers import NotificationSerializer
from notifications.models import Notification
from django.db import transaction
import logging


from cryptography.fernet import Fernet

key = b'Y2hhbmdlTWVOb3dLZXkxMjM0NTY3ODkwMTIzNDU2Nzg='  
cipher = Fernet(key)

def encrypt_user_id(user_id):
    return cipher.encrypt(str(user_id).encode()).decode()

def decrypt_user_id(encrypted_user_id):
    return int(cipher.decrypt(encrypted_user_id.encode()).decode())

def convert_duration(seconds):
    # Check if the value is None
    if seconds is None:
        return "N/A"  # Or another default value that makes sense for your case

    # If the input is a timedelta object, convert to total seconds
    if isinstance(seconds, timedelta):
        seconds = seconds.total_seconds()

    # If the duration is less than 60 seconds, show seconds
    if seconds < 60:
        return f"{int(seconds)} sec"

    # If the duration is greater than or equal to 60 seconds, convert to minutes
    elif seconds < 3600:  # 3600 seconds = 1 hour
        minutes = seconds // 60
        remaining_seconds = int(seconds % 60)
        return f"{int(minutes)} min" if remaining_seconds == 0 else f"{int(minutes)} min {remaining_seconds} sec"

    # If the duration is greater than or equal to 3600 seconds, convert to hours
    else:
        hours = seconds // 3600
        remaining_minutes = (seconds % 3600) // 60
        remaining_seconds = int(seconds % 60)
        if remaining_minutes == 0 and remaining_seconds == 0:
            return f"{int(hours)} hr"
        elif remaining_seconds == 0:
            return f"{int(hours)} hr {int(remaining_minutes)} min"
        else:
            return f"{int(hours)} hr {int(remaining_minutes)} min {remaining_seconds} sec"

def send_message(host, port, username, password, number, message):
    
    message = message.encode('utf-8')

    
    url = f"http://{host}:{port}/http/send-message/"
    params = {
        'username': username,
        'password': password,
        'to': number,
        'message': message,
    }

    try:
        
        response = requests.get(url, params=params)

       
        response.raise_for_status()

        
        return response.text
    except requests.RequestException as e:
        return f"Error: {e}"
#send_message('192.168.110.3', 9710, 'admin', 'admin123', phone_number, message)
class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer
@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def api_endpoints(request):
    endpoints = {
        "signup": "/signup/",
        "login": "/login/",
        "to get token":"/token/",
        "to refresh token":"/token/refresh/",
        "to get authenticated user":"/getuser/",
        "list_ticket": "/list_ticket/",
        "create_ticket": "/create_ticket/",
        "ticket": "/list_ticket/<int:pk>/",
        "edit_ticket": "/edit_ticket/<int:pk>/",
        "list_ticket_category": "/list_ticket_category/",
        "create_ticket_category": "/create_ticket_category/",
        "list_ticket_category_detail": "/list_ticket_category_detail/<int:pk>/",
        "delete_ticket_category": "/delete_ticket_category/<int:pk>/",
    }
    return Response(endpoints)
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
class GenerateAPIKeyView(APIView):
    def post(self, request):
        user = request.user
        api_key, created = APIKey.objects.get_or_create(user=user)
        if created:
            api_key.key = secrets.token_urlsafe(32)
            api_key.save()
        elif not api_key.key:
            api_key.key = secrets.token_urlsafe(32)
            api_key.save()
        return Response({'api_key': api_key.key})
@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def signup(request):
    serializer = UserSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        user = User.objects.get(username=request.data['username'])
        user.set_password(request.data['password'])
        user.save()
        token = Token.objects.create(user=user)
        return Response({'token': token.key, 'user': serializer.data})
    return Response(serializer.errors, status=status.HTTP_200_OK)
class GetUserView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        # Assuming you have a custom User model with a 'role' field
        # or a separate Role model with a foreign key to User
        role = user.role  # or user.userprofile.role, etc.
        return Response({'username': user.username, 'role': role})
@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def login(request):
    user = get_object_or_404(User, email=request.data['username'])
    if not user.check_password(request.data['password']):
        return Response("missing user", status=status.HTTP_404_NOT_FOUND)
    token, created = Token.objects.get_or_create(user=user)
    serializer = UserSerializer(user)
    response = Response({'token': token.key,'user': serializer.data})
    
    
    response.set_cookie(
    key='auth_token',
    value=token.key,
    path='/',
    httponly=True,         
    samesite='Lax'      # Change to 'Lax' or 'None' based on your requirements
)

    
    return response
@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def logout(request):
    # Retrieve the token from cookies
    token_key = request.COOKIES.get('auth_token')

    if token_key:
        try:
            token = Token.objects.get(key=token_key)
            token.delete()  # Delete the token to log out the user

            # Clear the cookie
            response = Response({'message': 'Logged out successfully.'}, status=status.HTTP_200_OK)
            response.delete_cookie('auth_token')  # Clear the cookie
            return response

        except Token.DoesNotExist:
            return Response({'error': 'Token does not exist.'}, status=status.HTTP_400_BAD_REQUEST)

    return Response({'error': 'User is not authenticated.'}, status=status.HTTP_401_UNAUTHORIZED)


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def GetUser(request):
    token = request.COOKIES.get('auth_token')
    print(f"Received token: {token}")  # Add debug logging

    if token:
        try:
            user = Token.objects.get(key=token).user
            serializer = UserGetSerializer(user)
            return Response(serializer.data)
        except Token.DoesNotExist:
            print("Token not found")  # Add debug logging
            return Response("Unauthorized", status=status.HTTP_401_UNAUTHORIZED)
    else:
        print("No token provided")  # Add debug logging
        return Response("Unauthorized", status=status.HTTP_401_UNAUTHORIZED)


class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer
class CustomTokenRefreshView(TokenRefreshView):
    serializer_class = CustomTokenRefreshSerializer
    
    
#views for ticket API
@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def create_ticket(request):
   
    serializer = TicketSerializer(data=request.data)
    
    if serializer.is_valid():
   
     ticket = serializer.save(created_by=request.user)

     files = request.FILES.getlist('attachments')
     for file in files:
        Attachment.objects.create(ticket=ticket, file=file)

     ticket_branch = ticket.created_by.branch
    
     it_officers = User.objects.filter(role="admin", branch=ticket_branch)

     for it_officer in it_officers:
        phone_number = it_officer.phone_number 

        
        requestor = request.user  
        ticket_title = ticket.title
        ticket_description = ticket.description
        ticket_status = ticket.status
        ticket_priority=ticket.priority
        ticket_category = ticket.category.name  
        assigned_to = ticket.assigned_to.first_name if ticket.assigned_to else "Not Assigned"
        requestor_name = f"{requestor.first_name} {requestor.last_name}"
        requestor_department = requestor.department  

 
        message = f"""
Awash Wine S.C IT Helpdesk
---------------------------
New Ticket Request

🔹 Requestor: {requestor_name}
🔹 Department: {requestor_department}
🔹 Category: {ticket_category}
🔹 Description: {ticket_description}
🔹 Priority: {ticket_priority}

---------------------------
Please review the details above and take the necessary actions as soon as possible.

For more information, log in to the IT Helpdesk system:
https://ithelpdesk.awashwines.info/

Best regards,
Awash Wine S.C IT Helpdesk
"""

        # Send the SMS
        print("Sending SMS to", phone_number)
        # send_message('drinkawash.com', 9797, 'admin', '@dIff%nSms0', phone_number, message)  # SMS sending function

        # Prepare email details for IT officers
        email_subject = 'New Support Ticket Request'
        ticketurl=f"https://ithelpdesk.awashwines.info/"
        # Render the email using the template (pass ticket info and IT officer data)
        email_message = render_to_string('new_ticket.html', {
            'user': ticket.created_by,
            'ticket_description': ticket_description,
            'ticket_created_at': ticket.created_at,
            'ticket_title': ticket_title,
            'assigned_to': assigned_to,
            'ticket_status': ticket_status,
            'ticket_category': ticket_category,
            'requestor_name': requestor_name,
            'requestor_department': requestor_department,
            'ticket_priority':ticket_priority,
            'ticketurl':ticketurl
        })
        print(f"sending email to {it_officer.email}")
        send_mail(
            email_subject,              # Subject of the email
            email_message,              # HTML email content
            settings.EMAIL_HOST_USER,   # Sender's email address (configured in Django settings)
            [it_officer.email],         # Recipient's email address (IT officer)
            html_message=email_message  # HTML content for the email
        )
        
    # Return a successful response
        return Response(serializer.data, status=status.HTTP_201_CREATED)

# If serializer is not valid, return errors
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@authentication_classes([APIKeyAuthentication])
@permission_classes([IsAuthenticated])
def list_tickets(request):
    tickets = Ticket.objects.all()
    serializer = TicketSerializer(tickets, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@authentication_classes([APIKeyAuthentication])
@permission_classes([IsAuthenticated])
def ticket(request, pk):
    tickets = Ticket.objects.get(id=pk)
    serializer = TicketSerializer(tickets)
    

    return Response(serializer.data)

@api_view(['POST'])
@authentication_classes([APIKeyAuthentication])
@permission_classes([IsAuthenticated])
def edit_ticket(request, pk):
    ticket=Ticket.objects.get(id=pk)
    serializer = TicketSerializer(ticket, data=request.data, partial=True)  # `partial=True` for partial updates
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



#views for ticket category
@api_view(['GET'])
@authentication_classes([APIKeyAuthentication])
@permission_classes([IsAuthenticated])
def list_ticket_category(request):
    ticket_category = TicketCategory.objects.all()
    serializer = TicketCategorySerializer(ticket_category, many=True)
    return Response(serializer.data)
    
@api_view(['POST'])
@authentication_classes([APIKeyAuthentication])
@permission_classes([IsAuthenticated])
def create_ticket_category(request):
    serializer = TicketCategorySerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
@api_view(['GET'])
@authentication_classes([APIKeyAuthentication])
@permission_classes([IsAuthenticated])
def list_ticket_category_detail(request,pk):
    ticket_category = TicketCategory.objects.get(id=pk)
    serializer = TicketCategorySerializer(ticket_category)
    return Response(serializer.data)

@api_view(['GET'])
@authentication_classes([APIKeyAuthentication])
@permission_classes([IsAuthenticated])
def departments(request):
    departments = Department.objects.all()
    serializer = DepartmentSerializer(departments, many=True)
    return Response(serializer.data)


@api_view(['POST'])
@authentication_classes([APIKeyAuthentication])
@permission_classes([IsAuthenticated])
def create_department(request):
    serializer = DepartmentSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PATCH'])
@authentication_classes([APIKeyAuthentication])
@permission_classes([IsAuthenticated])
def edit_department(request, pk):
    department=Department.objects.get(id=pk)
    serializer = DepartmentSerializer(department, data=request.data, partial=True)  # `partial=True` for partial updates
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@authentication_classes([APIKeyAuthentication])
@permission_classes([IsAuthenticated])

def delete_ticket_category(request,pk):
    ticket_category = TicketCategory.objects.get(id=pk)
    ticket_category.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)




## Admin API

@api_view(['GET'])
@authentication_classes([APIKeyAuthentication])
@permission_classes([IsAuthenticated])
def admin_dashboard(request):
    current_month = datetime.now().month
    # Count ticket status
    open_tickets = Ticket.objects.filter(status='Open').count()
    in_progress_tickets = Ticket.objects.filter(status='In Progress').count()
    closed_tickets = Ticket.objects.filter(status='Closed').count()
    pending_tickets=Ticket.objects.filter(status='Pending').count()

    


   
    lideta_requests = Ticket.objects.filter(created_by__branch='Lideta').count()
    mekanissa_requests = Ticket.objects.filter(created_by__branch='Mekanissa' ).count()
    farm_requests = Ticket.objects.filter(created_by__branch='Farm' ).count()

    # Get category counts
    categories = TicketCategory.objects.all()
    category_counts = {}
    for category in categories:
        category_counts[category.name] = Ticket.objects.filter(category=category).count()

    # Get recent 10 requests
    recent_requests = Ticket.objects.all().order_by('-created_at')[:10]
    recent_requests_serializer = RecentTicketSerializer(recent_requests, many=True)

    # Create response data
    data = {
        'ticket_status': {
            'open': open_tickets,
            'in_progress': in_progress_tickets,
            'closed': closed_tickets,
            'pending':pending_tickets
        },
        'branch_requests': {
            
            'Lideta': lideta_requests,
            'Mekanissa': mekanissa_requests,
            'Farm': farm_requests
        },
        'categories': category_counts,
        'recent_requests': recent_requests_serializer.data
    }

    return Response(data)



#users
@api_view(['GET'])
@authentication_classes([APIKeyAuthentication])
@permission_classes([IsAuthenticated])
def get_all_users(request):
    users = User.objects.all()  
    user_data = []
    
    for user in users:
        user_data.append({
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'phone_number': user.phone_number,
            'profile_picture': user.profile_picture.url if user.profile_picture else None,
            'branch': user.branch,
            'department': user.department.name if user.department else None, 
            'role': user.role,
            'first_name':user.first_name,
            'last_name':user.last_name,
            'is_active':user.is_active,
            
        })
    
    return Response(user_data, status=status.HTTP_200_OK)

#edit user
@api_view(['PATCH'])
@authentication_classes([APIKeyAuthentication])
@permission_classes([IsAuthenticated])
def edit_user(request, pk):
    try:
        user = User.objects.get(id=pk)  
    except User.DoesNotExist:
        return Response({'detail': 'User  not found.'}, status=status.HTTP_404_NOT_FOUND)

    
    serializer = UserSerializer(user, data=request.data, partial=True)  
    if serializer.is_valid():
        serializer.save()  
        return Response(serializer.data, status=status.HTTP_200_OK)

    
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
@api_view(['POST'])
@authentication_classes([APIKeyAuthentication])
@permission_classes([IsAuthenticated])
def create_user_signup(request):
    serializer = UserCreateSerializer(data=request.data)
    if serializer.is_valid():
        user=serializer.save()
        ActivateAccount.objects.create(user=user)

        encrypted_user_id = encrypt_user_id(user.id)

        
        activation_link = f"http://yourdomain.com/activate/{encrypted_user_id}/"

        
        
        email_subject = 'Activate Your Account'
        email_message = render_to_string('activation_email.html', {
            'user': user,
            'activation_link': activation_link,
        })

  
        send_mail(
            email_subject,
            email_message,
            settings.EMAIL_HOST_USER,
            [user.email],
            html_message=email_message 
        )

        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



@api_view(['POST'])
@authentication_classes([APIKeyAuthentication])
@permission_classes([IsAuthenticated])
def create_user(request):
    serializer = UserCreateSerializer(data=request.data)
    if serializer.is_valid():
        user=serializer.save()
        ActivateAccount.objects.create(user=user)

        encrypted_user_id = encrypt_user_id(user.id)

        
        activation_link = f"https://ithelpdesk.awashwines.info/activate/{encrypted_user_id}/"

        
        
        email_subject = 'Activate Your Account'
        email_message = render_to_string('activation_email.html', {
            'user': user,
            'activation_link': activation_link,
        })

  
        send_mail(
            email_subject,
            email_message,
            settings.EMAIL_HOST_USER,
            [user.email],
            html_message=email_message 
        )
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)




class ListNotificationsView(APIView):
    permission_classes = [IsAuthenticated]
    @authentication_classes([JWTAuthentication])

    def get(self, request):
        notifications = Notification.objects.filter(user=request.user,read=False).order_by('-created_at')
        serializer = NotificationSerializer(notifications, many=True)
        return Response(serializer.data)

class MarkNotificationAsReadView(APIView):
    permission_classes = [IsAuthenticated]
    @authentication_classes([JWTAuthentication])

    def post(self, request, pk):
        try:
            notification = Notification.objects.get(pk=pk, user=request.user)
            notification.read = True
            notification.save()
            return Response({'status': 'notification marked as read'}, status=status.HTTP_200_OK)
        except Notification.DoesNotExist:
            return Response({'error': 'Notification not found or does not belong to the user'}, status=status.HTTP_404_NOT_FOUND)
class MarkAllNotificationAsRead(APIView):
    permission_classes = [IsAuthenticated]
    @authentication_classes([JWTAuthentication])

    def post(self, request):
        try:
            notification = Notification.objects.filter( user=request.user)
            for notifications in notification:
             notifications.read = True
             notifications.save()
            return Response({'status': 'all notification marked as read'}, status=status.HTTP_200_OK)
        except Notification.DoesNotExist:
            return Response({'error': 'Notification not found or does not belong to the user'}, status=status.HTTP_404_NOT_FOUND)
    

@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def my_ticket(request):
    tickets = Ticket.objects.filter(created_by=request.user).order_by('-created_at')
    serializer = TicketSerializer(tickets, many=True)

    return Response(serializer.data)
@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def send_message_view(request):
    phone_number = request.data.get('phone_no')
    message = request.data.get('message')

    if not phone_number or not message:
        return Response({'error': 'Missing phone number or message'}, status=status.HTTP_400_BAD_REQUEST)

    send_message('192.168.110.11', 9710, 'admin', '@dIff%nSms0', phone_number, message)
    return Response({'status': 'Message sent successfully'}, status=status.HTTP_200_OK)
    
    
class SubmitSolutionView(APIView):
    permission_classes = [IsAuthenticated]
    @authentication_classes([JWTAuthentication])
    def post(self, request, ticket_id):
        try:
            ticket = Ticket.objects.get(id=ticket_id)
        except Ticket.DoesNotExist:
            return Response({'detail': 'Ticket not found.'}, status=status.HTTP_404_NOT_FOUND)
        if ticket.status == 'Closed':
            return Response({'detail': 'Ticket is already closed.'}, status=status.HTTP_400_BAD_REQUEST)
        data = {
            'ticket': ticket.id,
            'author': request.user.id,
            'content': request.data.get('content'),
            'parent': None  
        }
        serializer = TicketCommentSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
         
            
            tickets=Ticket.objects.get(id=ticket_id)
            tickets.status = 'Pending'
            tickets.save()
            
            Notification.objects.create(
            user=ticket.created_by,
            message=f"Your ticket '{ticket.title}' has been resolved by {ticket.assigned_to.get_full_name()}. Kindly close the ticket to confirm that the issue has been successfully addressed.",
            notification_type='Ticket'
                )
            


            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def ticket_status_history(request,id):
    
    try:
        ticket = Ticket.objects.get(id=id)
        
        try:
            TicketHistory.objects.get(ticket=ticket)
            ticket_history=TicketHistory.objects.get(ticket=ticket)
            new_value = request.data.get('new_value')
            old_value=request.data.get('old_value')
            field_updated=request.data.get('field_updated')
            TicketHistory.objects.create(
                ticket=ticket,
                updated_by=request.user,
                field_name=field_updated,
                new_value=new_value,
                old_value=old_value
            )
            ticket.save()
            return Response({'status': 'ticket history updated'}, status=status.HTTP_200_OK)
        except:
            old_value=request.data.get('old_value')
            new_value = request.data.get('new_value')
            field_updated=request.data.get('field_updated')
            TicketHistory.objects.create(
                ticket=ticket,
                updated_by=request.user,
                field_name=field_updated,
                new_value=new_value,
                old_value=old_value
            )
            ticket.save()
            
    
    
            return Response({'status': 'ticket history updated'}, status=status.HTTP_200_OK)
    except Ticket.DoesNotExist:
        return Response({'detail': 'Ticket not found'}, status=status.HTTP_404_NOT_FOUND)
    
    
@api_view(['GET'])
@authentication_classes([APIKeyAuthentication])
@permission_classes([IsAuthenticated])
def solutions(request,id):
    history = TicketHistory.objects.filter(ticket=id)
    serializer = TicketHistorySerializer(history, many=True)

    return Response(serializer.data)
    
    
logger = logging.getLogger(__name__)

@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def acceptticket(request, id):
    try:
        ticket = Ticket.objects.get(id=id)

      
        if ticket.assigned_to is not None:
            return Response({'error': 'Ticket is already assigned'}, status=status.HTTP_409_CONFLICT)


        open_tickets = Ticket.objects.filter(assigned_to=request.user, status__in=["In Progress", "Open"])
        if open_tickets.exists():
            return Response({'error': 'You have other open tickets. Please complete them before accepting a new ticket.'}, status=status.HTTP_403_FORBIDDEN)

        with transaction.atomic():
            ticket.assigned_to = request.user
            ticket.status = "In Progress"
            ticket.save()
        
        logger.info(f'Ticket {id} accepted by user {request.user.id}')
        requestor = request.user
        requestor_name = f"{requestor.first_name} {requestor.last_name}"
        phone=ticket.created_by
        requestor_message = f"""
Awash Wine S.C IT Helpdesk
---------------------------
Your Ticket Request has been successfully Accepted.
Your request is currently under review by {requestor_name}.

For updates, please log in to the IT Helpdesk system:
https://localhost:3000/

Best regards,
Awash Wine S.C IT Helpdesk
"""
        send_message('drinkawash.com', 9797, 'admin', '@dIff%nSms0', phone.phone_number, requestor_message)
        print("message sent")
        Notification.objects.create(
            user=ticket.created_by,
            message=f"Your ticket '{ticket.title}' has been accepted by {ticket.assigned_to.get_full_name()}.",
            notification_type='Ticket'
                )
        return Response({'status': 'Ticket accepted successfully', 'ticket_id': ticket.id, 'assigned_to': ticket.assigned_to.username}, status=status.HTTP_200_OK)

    except Ticket.DoesNotExist:
        return Response({'error': 'Ticket not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f'Error accepting ticket {id}: {str(e)}')
        return Response({'error': 'An error occurred while accepting the ticket'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def check_pending_tickets(request):
   
    has_pending_tickets = Ticket.objects.filter(status='Pending',created_by=request.user).exists()
    
    return Response({'has_pending_tickets': has_pending_tickets})
@api_view(['POST'])
@authentication_classes([APIKeyAuthentication])
@permission_classes([IsAuthenticated])
def close_ticket(request,id):
    ticket=Ticket.objects.get(id=id)
    ticket.status="Closed"
    ticket.save()
    Notification.objects.create(
            user=ticket.assigned_to,
            message=f"Ticket '{ticket.title}' has been closed  by {ticket.created_by.get_full_name()}.",
            notification_type='Ticket'
                )
    return Response({'status': 'ticket closed '}, status=status.HTTP_200_OK)

@api_view(['GET'])
@authentication_classes([APIKeyAuthentication])
@permission_classes([IsAuthenticated])
def list_solution(request,id):
    solutions=TicketComment.objects.filter(ticket=id)
    serializer = TicketCommentSerializer(solutions, many=True)

    return Response(serializer.data)


@api_view(['GET'])
@authentication_classes([APIKeyAuthentication])
@permission_classes([IsAuthenticated])
def DownloadAttachmentView(request, id):
    attachment = get_object_or_404(Attachment, ticket=id)
    if not attachment.file:
        raise Http404("Attachment not found")
    response = FileResponse(attachment.file, as_attachment=True, filename=attachment.file.name)
    return response

@api_view(['POST'])
@authentication_classes([APIKeyAuthentication])
@permission_classes([IsAuthenticated])
def password_reset(request):
    email = request.data.get('email')

    if not email:
        return Response({"error": "Email is required"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        user = User.objects.get(email=email)
        password_reset_entry, created = PasswordReset.objects.get_or_create(user=user)


        if not created:

            password_reset_entry.token = str(uuid.uuid4()) 
            password_reset_entry.save() 
        

        reset_url = f"https://ithelpdesk.awashwines.info/reset/{urlsafe_base64_encode(str(user.pk).encode())}/{password_reset_entry.token}/"


        email_subject = 'Password Reset Request'
        email_message = render_to_string('password_reset_email.html', {
            'user': user,
            'reset_url': reset_url
        })

  
        send_mail(
            email_subject,
            email_message,
            settings.EMAIL_HOST_USER,
            [user.email],
            html_message=email_message 
        )

        return Response({"message": "Password reset email sent."}, status=status.HTTP_200_OK)

    except User.DoesNotExist:
        return Response({"error": "Email not found"}, status=status.HTTP_404_NOT_FOUND)
    
@api_view(['POST'])
@authentication_classes([APIKeyAuthentication])
@permission_classes([IsAuthenticated])
def change_password(request, uidb64, token):
    """
    API view to allow user to reset their password using a valid reset token.
    """

    try:
 
        uid = urlsafe_base64_decode(uidb64).decode()
        user = get_user_model().objects.get(pk=uid)


        password_reset_entry = PasswordReset.objects.filter(user=user, token=token).first()


        if not password_reset_entry:
            return Response({"error": "Invalid token."}, status=status.HTTP_400_BAD_REQUEST)


        new_password = request.data.get('password')

        if not new_password:
            return Response({"error": "Password is required."}, status=status.HTTP_400_BAD_REQUEST)


        user.set_password(new_password)
        user.save()

      
        password_reset_entry.delete()

        return Response({"message": "Password has been successfully reset."}, status=status.HTTP_200_OK)

    except User.DoesNotExist:
        return Response({"error": "User not found."}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
    

@api_view(['GET'])
@authentication_classes([APIKeyAuthentication])
@permission_classes([IsAuthenticated])
def TicketReportView(request):

    start_date_filter = request.GET.get('start_date', None)
    end_date_filter = request.GET.get('end_date', None)
    status_filter = request.GET.get('status', None)
    branch_filter = request.GET.get('branch', None)
    department_filter = request.GET.get('department', None)
    assigned_to_filter = request.GET.get('assigned_to', None)
    category_filter = request.GET.get('category', None)
    export_flag = request.GET.get('export', 'false')


    filters = Q()

    if start_date_filter:
        start_date = parse_datetime(start_date_filter)
        if start_date:
    
            if timezone.is_naive(start_date):
                start_date = timezone.make_aware(start_date, timezone.get_current_timezone())
            filters &= Q(created_at__gte=start_date)

    if end_date_filter:
        end_date = parse_datetime(end_date_filter)
        if end_date:

            end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)
            if timezone.is_naive(end_date):
                end_date = timezone.make_aware(end_date, timezone.get_current_timezone())
            filters &= Q(created_at__lte=end_date)

    if status_filter:
        filters &= Q(status=status_filter)

    if branch_filter:
        filters &= Q(created_by__branch=branch_filter)

    if department_filter:
        filters &= Q(created_by__department__name=department_filter)

    if assigned_to_filter:
        filters &= Q(assigned_to__email=assigned_to_filter)

    if category_filter:
        filters &= Q(category__name=category_filter)

    tickets = Ticket.objects.filter(filters).order_by('-created_at')


    


    ticket_serializer = ReportTicketSerializer(tickets, many=True)

    ticket_status_counts = tickets.values('status') \
        .annotate(total=Count('id')) \
        .order_by('status')

    department_counts = tickets.values('created_by__department__name') \
        .annotate(total=Count('id')) \
        .order_by('created_by__department__name')

    category_counts = tickets.values('category__name') \
        .annotate(total=Count('id')) \
        .order_by('category__name')

    branch_counts = tickets.values('created_by__branch') \
        .annotate(total=Count('id')) \
        .order_by('created_by__branch')

    case_holder_counts = tickets.values('assigned_to__first_name', 'assigned_to__last_name') \
    .annotate(
        assigned_to__=Concat(
            F('assigned_to__first_name'), 
            Value(' '), 
            F('assigned_to__last_name')
        )
    ) \
    .annotate(total=Count('id')) \
    .order_by('assigned_to__')

    try:
        avg_response_time = TicketHistory.objects.filter(field_name='status', new_value='In Progress', ticket__in=tickets) \
            .values('ticket') \
            .annotate(
                response_time=Avg(
                    ExpressionWrapper(
                        F('updated_at') - F('ticket__created_at'),
                        output_field=fields.DurationField()
                    )
                )
            ).aggregate(Avg('response_time'))
    except:
        avg_response_time=0
    
    try:
        avg_fixing_time = TicketHistory.objects.filter(field_name='status', new_value='Pending', ticket__in=tickets) \
            .values('ticket') \
            .annotate(
                fixing_time=Avg(
                    ExpressionWrapper(
                        F('updated_at') - F('ticket__created_at'),
                        output_field=fields.DurationField()
                    )
                )
            ).aggregate(Avg('fixing_time'))
    except:
        avg_fixing_time=0
    
    statistics = {
        'ticket_status_counts': ticket_status_counts,
        'department_counts': department_counts,
        'category_counts': category_counts,
        'branch_counts': branch_counts,
        'case_holder_counts': case_holder_counts,
        'avg_response_time': convert_duration(avg_response_time['response_time__avg']),
        'avg_fixing_time': convert_duration(avg_fixing_time['fixing_time__avg'])
    }

    response_data = {
        'report': ticket_serializer.data,
        'statistics': statistics
    }
    if export_flag.lower() == 'true':
        return generate_excel_report(tickets, avg_response_time, avg_fixing_time)

    return Response(response_data, status=status.HTTP_200_OK)
def generate_excel_report(tickets, avg_response_time, avg_fixing_time):
    data = list(tickets.values(
        'id', 'status', 'created_by__branch', 'created_at',
        'created_by__department__name', 'assigned_to__first_name',
        'category__name', 'created_by__first_name', 'created_by__last_name'
    ))

    df = pd.DataFrame(data)
    df['created_at'] = df['created_at'].apply(lambda x: x.replace(tzinfo=None))
    df['Created By'] = df['created_by__first_name'] + ' ' + df['created_by__last_name']
    df.drop(['created_by__first_name', 'created_by__last_name'], axis=1, inplace=True)
    df.columns = ['Ticket ID', 'Status', 'Branch', 'Created At', 'Department', 'Assigned To', 'Category', 'Created By']
    df.reset_index(drop=True, inplace=True)
    df = df[['Ticket ID', 'Status', 'Branch', 'Created At', 'Department', 'Assigned To', 'Category', 'Created By']]
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Ticket Report"

    for row in dataframe_to_rows(df, index=False, header=True):
        ws1.append(row)

    
    ws2 = wb.create_sheet(title="Statistics")

    ticket_status_counts = df['Status'].value_counts().to_dict()
    department_counts = df['Department'].value_counts().to_dict()
    category_counts = df['Category'].value_counts().to_dict()
    branch_counts = df['Branch'].value_counts().to_dict()
    assigned_to_counts = df['Assigned To'].value_counts().to_dict()

    ws2.append(['Statistic', 'Count'])

    ws2.append(['Ticket Status Counts'])
    for status, count in ticket_status_counts.items():
        ws2.append([status, count])

    ws2.append([''])  

    ws2.append(['Department Counts'])
    for department, count in department_counts.items():
        ws2.append([department, count])

    ws2.append([''])  

    ws2.append(['Category Counts'])
    for category, count in category_counts.items():
        ws2.append([category, count])

    ws2.append([''])  

    ws2.append(['Branch Counts'])
    for branch, count in branch_counts.items():
        ws2.append([branch, count])

    ws2.append([''])  

    ws2.append(['Assigned To Counts'])
    for assignee, count in assigned_to_counts.items():
        ws2.append([assignee, count])

    ws2.append([''])  

    avg_response_time_delta = avg_response_time.get('response_time__avg')
    avg_fixing_time_delta = avg_fixing_time.get('fixing_time__avg')

    if avg_response_time_delta is not None:
        avg_response_time_seconds = avg_response_time_delta.total_seconds()
    else:
        avg_response_time_seconds = 0

    if avg_fixing_time_delta is not None:
        avg_fixing_time_seconds = avg_fixing_time_delta.total_seconds()
    else:
        avg_fixing_time_seconds = 0


    avg_response_time_formatted = convert_duration(avg_response_time_seconds)
    avg_fixing_time_formatted = convert_duration(avg_fixing_time_seconds)

   
    ws2.append(['Avg Response Time', avg_response_time_formatted])
    ws2.append(['Avg Fixing Time', avg_fixing_time_formatted])

   
    from openpyxl.styles import Font, Alignment

    
    for cell in ws2[1]:
        cell.font = Font(bold=True)

   
    for row in ws2.iter_rows(min_row=2, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

   
    file_stream = BytesIO()
    wb.save(file_stream)

    
    file_stream.seek(0)

    
    response = HttpResponse(file_stream, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ticket_report.xlsx"'

    return response
    
    
@api_view(['GET'])
@authentication_classes([APIKeyAuthentication])
@permission_classes([IsAuthenticated])
def check_activation_status(request, user_id):
    try:
        user = get_object_or_404(User, id=user_id)
        activation_record = ActivateAccount.objects.get(user=user)
        

        return Response({"activated": activation_record.activated}, status=status.HTTP_200_OK)
    except ActivateAccount.DoesNotExist:
        return Response({"activated": False}, status=status.HTTP_404_NOT_FOUND)

@api_view(['POST'])
def activate_account(request, encrypted_user_id):
    try:
    
        user_id = decrypt_user_id(encrypted_user_id)

   
        user = get_object_or_404(User, id=user_id)

    
        activation_record, created = ActivateAccount.objects.get_or_create(user=user)

        if activation_record.activated:
            return Response({"message": "Your account is already activated."}, status=status.HTTP_400_BAD_REQUEST)

        activation_record.activated = True
        activation_record.save()

        return Response({"message": "Your account has been activated successfully!"}, status=status.HTTP_200_OK)

    except ValueError:

        return Response({"error": "Invalid activation link."}, status=status.HTTP_400_BAD_REQUEST)
    except User.DoesNotExist:

        return Response({"error": "User  does not exist."}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
    
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
@api_view(['GET'])
@authentication_classes([APIKeyAuthentication])
@permission_classes([IsAuthenticated])
def acknowledgement(request, ticket_id):
    try:
        acknowledgements = Acknowledgement.objects.filter(ticket=ticket_id)
        serializer = AcknowledgementSerializer(acknowledgements, many=True)
        return Response(serializer.data)
    except Acknowledgement.DoesNotExist:
        return Response({'error': 'Acknowledgement not found'}, status=status.HTTP_404_NOT_FOUND)

@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def create_acknowledgement(request):
    serializer = AcknowledgementSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save(author=request.user)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@authentication_classes([APIKeyAuthentication])
@permission_classes([IsAuthenticated])
def list_acknowledgements(request):
    acknowledgements = Acknowledgement.objects.all()
    serializer = AcknowledgementSerializer(acknowledgements, many=True)
    return Response(serializer.data)

@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def reverse_ticket(request, ticket_id):
    try:
       
        ticket = Ticket.objects.get(id=ticket_id)

        
        if ticket.status == 'In Progress':
            return Response({"detail": "Ticket is already in progress."}, status=status.HTTP_400_BAD_REQUEST)

       
        TicketHistory.objects.create(
            ticket=ticket,
            updated_by=request.user, 
            field_name='status',
            old_value=ticket.status,
            new_value='Reversed'
        )

        
        ticket.status = 'In Progress'
        ticket.save()

        return Response({"detail": "Ticket status updated to 'In Progress'."}, status=status.HTTP_200_OK)
    
    except Ticket.DoesNotExist:
        return Response({"detail": "Ticket not found."}, status=status.HTTP_404_NOT_FOUND)